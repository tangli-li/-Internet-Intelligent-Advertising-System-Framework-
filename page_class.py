import tkinter as tk
from tkinter import ttk
import pandas as pd
import random
from openpyxl import load_workbook
from choose_ad import User_coming

paths=r".\data\ad_system.xlsx"
#AD_display=pd.read_excel(paths,sheet_name=0)
#USER=pd.read_excel(paths,sheet_name=1)
#AD_click=pd.read_excel(paths,sheet_name=2)
#ADHOST=pd.read_excel(paths,sheet_name=3)

class AD_main(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        self.ID=int(ID)
        ADHOST = pd.read_excel(paths, sheet_name=3)
        total=ADHOST.loc[ADHOST["广告主id"]==self.ID,"广告总预算"].to_numpy()
        print(total)
        rest=ADHOST.loc[ADHOST["广告主id"]==self.ID,"广告预算剩余"].to_numpy()

        style = ttk.Style()
        style.configure('TCombobox', padding=5)
        style.configure('TButton', padding=5)

        lbl_visitor = ttk.Label(self, text="访问者类型:广告主", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        # 中间区域：预算信息容器
        frame_center = ttk.Frame(self)
        frame_center.grid(row=1, column=1, sticky='nsew')

        # 总预算标签
        lbl_total = ttk.Label(frame_center,
                              text=f"广告总预算：{float(total)}元",
                              font=('Arial', 12))
        lbl_total.pack(pady=10)

        # 剩余预算标签
        lbl_remaining = ttk.Label(frame_center,
                                  text=f"广告预算剩余：{float(rest)}元",
                                  font=('Arial', 12),
                                  foreground='blue')
        lbl_remaining.pack(pady=10)

        # 右下角：返回按钮
        btn_back = ttk.Button(self,
                              text="返回",
                              command=lambda: controller.create_newpage(login,None))
        btn_back.grid(row=2, column=2, padx=10, pady=10, sticky='se')

        proceed = ttk.Button(self,
                              text="确定",
                              command=lambda: controller.create_newpage(playad_page,self.ID))
        proceed.grid(row=2, column=1, padx=10, pady=10, sticky='se')

class Usr_page(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        self.ID=int(ID)
        USER = pd.read_excel(paths, sheet_name=1)
        def step():#填入搜索内容
            content=search.get()
            USER.loc[USER["用户id"]==self.ID,"搜索内容"]=content
            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                USER.to_excel(writer,sheet_name="用户信息表",index=False,header=True)
            controller.create_newpage(AD_show, self.ID)
        style = ttk.Style()
        style.configure('TButton', padding=3)

        lbl_visitor = ttk.Label(self, text="访问者类型:用户", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        frame=ttk.Frame(self)
        frame.grid(row=1,column=0,padx=10, pady=10, sticky='nw')

        ttk.Label(frame, text="搜索框").grid(row=1, column=0, sticky=tk.W, pady=5)
        search=ttk.Entry(frame,width=30)
        search.grid(row=1, column=1, padx=10, pady=5)

        btn_back = ttk.Button(self,
                              text="返回",
                              command=lambda: controller.create_newpage(login,None))
        btn_back.grid(row=2, column=2, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="确定",
                             command=step)
        proceed.grid(row=2, column=1, padx=10, pady=10, sticky=tk.SE)
#********************************************************************************
class AD_show(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        self.ID=int(ID)
        AD_display=pd.read_excel(paths,sheet_name=0)
        USER=pd.read_excel(paths,sheet_name=1)
        ADHOST=pd.read_excel(paths,sheet_name=3)
        AD_click = pd.read_excel(paths, sheet_name=2)

        def record(ad_id):
            new_row=pd.DataFrame(columns=AD_click.columns)
            usr_id=self.ID
            adhost_id=AD_display.loc[AD_display["广告id"]==ad_id,"广告主id"].to_list()
            #记录广告点击情况
            new_row.loc[0,"广告id"]=ad_id
            new_row.loc[0,"用户id"]=usr_id
            new_item0 = pd.concat([pd.DataFrame(columns=AD_click.columns), new_row], ignore_index=True)
            print(new_item0)
            book = load_workbook(paths)
            ws = book["广告点击表"]  # 确保工作表存在

            # 获取当前最大行号（从0开始计数）
            max_row = ws.max_row  # 正确获取最后一行的行号

            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                # 按原表列顺序重排数据
                existing_columns = [cell.value for cell in ws[1]]  # 假设第一行是标题
                s4_ordered = new_item0[existing_columns]  # 按原列顺序排序

                # 安全写入数据
                s4_ordered.to_excel(
                    writer,
                    sheet_name="广告点击表",
                    startrow=max_row,  # 从最后一行之后开始
                    index=False,
                    header=False
                )
            #广告点击后扣费
            ad_rest=ADHOST.loc[ADHOST["广告主id"]==adhost_id[0],"广告预算剩余"].to_list()
            price=AD_display.loc[AD_display["广告id"]==ad_id,"广告竞价"].to_list()
            rest_out=ad_rest[0]-price[0]
            ADHOST.loc[ADHOST["广告主id"] == adhost_id[0], "广告预算剩余"] = rest_out
            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                ADHOST.to_excel(writer, sheet_name="广告主信息表", index=False, header=True)

        User_search=USER.loc[USER["用户id"]==self.ID,"搜索内容"].to_list()
        person=User_coming(self.ID,User_search[0],AD_display,USER,ADHOST)
        ad_id_all,ad_content=person.run_all()

        lbl_visitor = ttk.Label(self, text="访问者类型:用户", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        frame = ttk.Frame(self)
        frame.grid(row=1, column=0, padx=10, pady=10, sticky='nw')

        ttk.Label(frame,text="广告1").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Label(frame, text="广告2").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Label(frame, text="广告3").grid(row=3, column=0, sticky=tk.W, pady=5)
        print(ad_content)

        ttk.Label(frame, text=ad_content[0]).grid(row=1, column=1, sticky=tk.W, pady=5)
        ttk.Label(frame, text=ad_content[1]).grid(row=2, column=1, sticky=tk.W, pady=5)
        ttk.Label(frame, text=ad_content[2]).grid(row=3, column=1, sticky=tk.W, pady=5)

        click1=ttk.Button(frame,text="点击",command=lambda: record(int(ad_id_all[0])))
        click1.grid(row=1, column=2, padx=10, pady=10)

        click2 = ttk.Button(frame, text="点击", command=lambda: record(int(ad_id_all[1])))
        click2.grid(row=2, column=2, padx=10, pady=10)

        click3 = ttk.Button(frame, text="点击", command=lambda: record(int(ad_id_all[2])))
        click3.grid(row=3, column=2, padx=10, pady=10)

        back=ttk.Button(self,text="退出",
                        command=lambda :controller.create_newpage(login,None))
        back.grid(row=2, column=2, padx=10, pady=10, sticky=tk.SE)

    def get_click(self):
        click1=0
        click2=0
        click3=0
        return [click1,click2,click3]

class playad_page(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        AD_display = pd.read_excel(paths, sheet_name=0)
        def input0():
            new_item.loc[0,need_combobox.get()]=need_entry.get()

        def step():
            new_item.loc[0, "广告竞价"] = price_entry.get()
            new_item.loc[0, "广告内容"] = content_entry.get()
            new_item.fillna("NA")
            new_item0=pd.concat([pd.DataFrame(columns=AD_display.columns), new_item], ignore_index=True)
            print(new_item0)
            book = load_workbook(paths)
            ws = book["广告投放表"]  # 确保工作表存在

            # 获取当前最大行号（从0开始计数）
            max_row = ws.max_row  # 正确获取最后一行的行号

            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                # 按原表列顺序重排数据
                existing_columns = [cell.value for cell in ws[1]]  # 假设第一行是标题
                s4_ordered = new_item0[existing_columns]  # 按原列顺序排序

                # 安全写入数据
                s4_ordered.to_excel(
                    writer,
                    sheet_name="广告投放表",
                    startrow=max_row,  # 从最后一行之后开始
                    index=False,
                    header=False
                )
                controller.create_newpage(playad_call,self.ID)

        self.ID=int(ID)
        new_item=pd.DataFrame(columns=AD_display.columns)

        style = ttk.Style()
        style.configure('TCombobox', padding=3)
        style.configure('TButton', padding=3)

        lbl_visitor = ttk.Label(self, text="访问者类型:广告主", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        AD_id = AD_display.loc[:, "广告id"].to_numpy()
        AD_list=map(str,AD_id)
        id="0000"
        id0 = str(self.ID)+"0000"
        while id0 in AD_list:
            id = str(random.randint(1e3, 1e4))
        AD_ID=str(self.ID)+id

        lbl_ad = ttk.Label(self, text=f"广告：{AD_ID}", font=('Arial', 14, 'bold'))
        lbl_ad.grid(row=0, column=2, padx=10, pady=10, sticky='nw')

        frame_center_1 = ttk.Frame(self)
        frame_center_1.grid(row=1, column=0, sticky='nsew')

        frame_center_2=ttk.Frame(self)
        frame_center_2.grid(row=1, column=2, sticky='nsew')

        list0=AD_display.columns.to_numpy()
        list=[]
        for i in range(4,len(list0)):
            list.append(list0[i])
        print(list)
        ttk.Label(frame_center_1, text="广告投放需求类型:").grid(row=1, column=0, sticky=tk.W, pady=5)
        need_combobox = ttk.Combobox(
            frame_center_1,
            values=list,
            state="readonly",
            width=10
        )
        need_combobox.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(frame_center_1,text="广告投放需求输入").grid(row=2,column=0,sticky=tk.W, pady=5)
        need_entry = ttk.Entry(frame_center_1, width=10)
        need_entry.grid(row=2, column=1, padx=10, pady=5)

        need_button=ttk.Button(frame_center_1,text="确认录入",command=input0)
        need_button.grid(row=3, column=1, padx=10, pady=10, sticky=tk.SE)

        ttk.Label(frame_center_2, text="广告内容输入").grid(row=1, column=2, sticky=tk.W, pady=5)
        content_entry = ttk.Entry(frame_center_2, width=10)
        content_entry.grid(row=1, column=3, padx=10, pady=5)

        ttk.Label(frame_center_2, text="竞价输入").grid(row=2, column=2, sticky=tk.W, pady=5)
        price_entry = ttk.Entry(frame_center_2, width=10)
        price_entry.grid(row=2, column=3, padx=10, pady=5)

        btn_back = ttk.Button(self,
                              text="返回",
                              command=lambda: controller.show_frame(AD_main))
        btn_back.grid(row=2, column=2, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="确定",
                             command=step)
        proceed.grid(row=2, column=1, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="增加",
                             command=lambda: "")
        proceed.grid(row=2, column=0, padx=10, pady=10, sticky=tk.SE)

        new_item.loc[0, "广告主id"] = self.ID
        new_item.loc[0, "广告id"] = AD_ID

class playad_call(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        self.ID=int(ID)
        lbl_visitor = ttk.Label(self, text="访问者类型:广告主", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        Call=ttk.Label(self,text="是否继续录入",font=('Arial', 16, 'bold'))
        Call.grid(row=1,column=1,padx=15, pady=15, sticky='nw')

        yes=ttk.Button(self, text="是",
                       command=lambda :controller.create_newpage(playad_page,self.ID))
        yes.grid(row=2, column=1, padx=10, pady=10, sticky="se")

        no = ttk.Button(self, text="退出",
                         command=lambda: controller.create_newpage(login,self))
        no.grid(row=2, column=2, padx=10, pady=10, sticky="se")

class visitor_register_usr(tk.Frame):
    def __init__(self,parent,controller,ID):
        super().__init__(parent)
        USER = pd.read_excel(paths, sheet_name=1)
        def input0():
            new_item.loc[0,need_combobox.get()]=need_entry.get()
        def step():
            new_item.fillna("NA")
            new_item0=pd.concat([pd.DataFrame(columns=USER.columns), new_item], ignore_index=True)
            print(new_item0)
            book = load_workbook(paths)
            ws = book["用户信息表"]  # 确保工作表存在

            # 获取当前最大行号（从0开始计数）
            max_row = ws.max_row  # 正确获取最后一行的行号

            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                # 按原表列顺序重排数据
                existing_columns = [cell.value for cell in ws[1]]  # 假设第一行是标题
                s4_ordered = new_item0[existing_columns]  # 按原列顺序排序

                # 安全写入数据
                s4_ordered.to_excel(
                    writer,
                    sheet_name="用户信息表",
                    startrow=max_row,  # 从最后一行之后开始
                    index=False,
                    header=False
                )
                controller.create_newpage(Usr_page,self.ID)

        self.ID=int(ID)
        new_item = pd.DataFrame(columns=USER.columns)
        new_item.loc[0,"用户id"]=self.ID

        lbl_visitor=ttk.Label(self,text="用户身份新建", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        frame_center_1 = ttk.Frame(self)
        frame_center_1.grid(row=1, column=0, sticky='nsew')

        list0 = USER.columns.to_numpy()
        list = []
        for i in range(2, len(list0)):
            list.append(list0[i])
        print(list)
        ttk.Label(frame_center_1, text="基础信息选择:").grid(row=1, column=0, sticky=tk.W, pady=5)
        need_combobox = ttk.Combobox(
            frame_center_1,
            values=list,
            state="readonly",
            width=10
        )
        need_combobox.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(frame_center_1, text="基础信息输入").grid(row=2, column=0, sticky=tk.W, pady=5)
        need_entry = ttk.Entry(frame_center_1, width=10)
        need_entry.grid(row=2, column=1, padx=10, pady=5)

        need_button = ttk.Button(frame_center_1, text="确认录入", command=input0)
        need_button.grid(row=3, column=1, padx=10, pady=10, sticky=tk.SE)

        btn_back = ttk.Button(self,
                              text="返回",
                              command=lambda: controller.create_newpage(login,None))
        btn_back.grid(row=2, column=2, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="确定",
                             command=step)
        proceed.grid(row=2, column=1, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="增加",
                             command=lambda: "")
        proceed.grid(row=2, column=0, padx=10, pady=10, sticky=tk.SE)


class visitor_register_adhost(tk.Frame):
    def __init__(self, parent, controller,ID):
        super().__init__(parent)
        ADHOST = pd.read_excel(paths, sheet_name=3)
        def step():
            new_item.loc[0,"广告总预算"]=ad_all.get()
            new_item.loc[0,"广告预算剩余"]=ad_all.get()
            new_item.loc[0,"历史广告被点击率"]=1
            new_item.fillna("NA")
            new_item0 = pd.concat([pd.DataFrame(columns=ADHOST.columns), new_item], ignore_index=True)

            book = load_workbook(paths)
            ws = book["广告主信息表"]  # 确保工作表存在

            # 获取当前最大行号（从0开始计数）
            max_row = ws.max_row  # 正确获取最后一行的行号

            with pd.ExcelWriter(
                    paths,
                    mode="a",  # 追加模式
                    engine="openpyxl",  # 必须指定引擎
                    if_sheet_exists="overlay"  # 覆盖写入模式
            ) as writer:
                # 按原表列顺序重排数据
                existing_columns = [cell.value for cell in ws[1]]  # 假设第一行是标题
                print(existing_columns)
                s4_ordered = new_item0[existing_columns]  # 按原列顺序排序

                # 安全写入数据
                s4_ordered.to_excel(
                    writer,
                    sheet_name="广告主信息表",
                    startrow=max_row,  # 从最后一行之后开始
                    index=False,
                    header=False
                )
            controller.create_newpage(AD_main,self.ID)

        self.ID = int(ID)
        new_item = pd.DataFrame(columns=ADHOST.columns)
        new_item.loc[0, "广告主id"] = self.ID

        lbl_visitor = ttk.Label(self, text="广告主身份新建", font=('Arial', 14, 'bold'))
        lbl_visitor.grid(row=0, column=0, padx=10, pady=10, sticky='nw')

        frame_center_1 = ttk.Frame(self)
        frame_center_1.grid(row=1, column=0, sticky='nsew')

        ttk.Label(frame_center_1,text="广告总预算投入").grid(row=1, column=0, sticky=tk.W, pady=5)
        ad_all=ttk.Entry(frame_center_1, width=10)
        ad_all.grid(row=1,column=1,padx=10, pady=5)

        btn_back = ttk.Button(self,
                              text="返回",
                              command=lambda: controller.create_newpage(login,None))
        btn_back.grid(row=2, column=2, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="确定",
                             command=step)
        proceed.grid(row=2, column=1, padx=10, pady=10, sticky=tk.SE)

        proceed = ttk.Button(self,
                             text="增加",
                             command=lambda: "")
        proceed.grid(row=2, column=0, padx=10, pady=10, sticky=tk.SE)

# class register(tk.Frame):
#     def __init__(self,parent,controller):
#         super().__init__(parent)
#         style = ttk.Style()
#         style.configure('TCombobox', padding=5)
#         style.configure('TButton', padding=5)
#
#         choose_label=ttk.Label(self,text="选择身份：",font=('Arial', 14, 'bold'))
#         choose_label.grid(row=0,column=0,padx=10,pady=10)
#
#         usr_button=ttk.Button(self,text="用户",
#                               command=controller.create_newpage)

class login(tk.Frame):
    def __init__(self,parent,controller,ID=None):
        super().__init__(parent)
        ADHOST = pd.read_excel(paths, sheet_name=3)
        USER = pd.read_excel(paths, sheet_name=1)
        # def register():#建议重新写一个页面
        #     id_vis = visitor_type_combobox.get()
        #     ID = user_id_entry.get()
        #     if id_vis=="广告主":
        #         controller.create_newpage(visitor_register_adhost, ID)
        #     if id_vis == "用户":
        #         controller.create_newpage(visitor_register_usr, ID)
        def click():
            id_vis=visitor_type_combobox.get()
            ID = user_id_entry.get()
            if id_vis=="广告主":
                List=ADHOST.loc[:,"广告主id"].to_numpy()
                print(List)
                if int(ID) in List:
                    controller.create_newpage(AD_main, ID)
                else:
                    controller.create_newpage(visitor_register_adhost, ID)
            if id_vis == "用户":
                List=USER.loc[:,"用户id"].to_numpy()
                if int(ID) in List:
                    controller.create_newpage(Usr_page,ID)
                else:
                    controller.create_newpage(visitor_register_usr, ID)

        style = ttk.Style()
        style.configure('TCombobox', padding=5)
        style.configure('TButton', padding=5)


        ttk.Label(
            self,
            text="互联网广告系统",
            font=("微软雅黑", 16, "bold"),
            foreground="#2c3e50"
        ).pack()

        main_frame = ttk.Frame(self)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20)

        # 输入字段布局
        input_frame = ttk.Frame(main_frame)
        input_frame.grid(row=0, column=0, sticky=tk.W)

        # 访问者类型
        ttk.Label(input_frame, text="访问者类型:").grid(row=0, column=0, sticky=tk.W, pady=5)
        visitor_type_combobox = ttk.Combobox(
            input_frame,
            values=["广告主","用户"],
            state="readonly",
            width=18
        )
        visitor_type_combobox.grid(row=0, column=1, padx=10, pady=5)

        # 用户ID
        ttk.Label(input_frame, text="用户ID:").grid(row=1, column=0, sticky=tk.W, pady=5)
        user_id_entry = ttk.Entry(input_frame, width=22)
        user_id_entry.grid(row=1, column=1, padx=10, pady=5)

        # 按钮容器（右下角定位）
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, sticky=tk.SE, pady=10)

        button_frame1 = ttk.Frame(main_frame)
        button_frame1.grid(row=1, column=0, sticky=tk.SE, pady=10)

        # 新建用户按钮
        # new_user_btn = ttk.Button(
        #     button_frame,
        #     text="新建用户",
        #     command=register,
        #     style="TButton"
        # )

        button0 = ttk.Button(
            button_frame,
            text="确定",
            command=click,
            style="TButton"
        )
        button0.pack(side=tk.LEFT, padx=5)

        # new_user_btn.pack(side=tk.RIGHT, padx=5)

        # 响应式布局配置
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
